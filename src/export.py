import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "output/inventory_clean.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(
    start_color="2F5496", end_color="2F5496", fill_type="solid"
)
HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)
SECTION_FONT = Font(bold=True, size=12)

CURRENCY_FMT = "#,##0.00"
INTEGER_FMT = "#,##0"


def _style_sheet(ws):
    """Aplica headers azules con texto blanco y ajusta ancho de columnas."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for col_idx, col_cells in enumerate(ws.columns, 1):
        sample = col_cells[:100]
        width = max(len(str(c.value or "")) for c in sample) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 30)


def _apply_number_format(ws, df, col_formats):
    """Aplica formato numerico a columnas del worksheet.

    Args:
        ws: Worksheet de openpyxl.
        df: DataFrame con las mismas columnas que el worksheet.
        col_formats: Dict {nombre_columna: formato_str}.
            Ejemplo: {"Price": "#,##0.00", "Qty": "#,##0"}
    """
    for col_name, fmt in col_formats.items():
        if col_name not in df.columns:
            continue
        col_idx = df.columns.get_loc(col_name) + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = fmt


def _build_summary(inv, orders):
    """Construye hoja Summary con KPIs de negocio y cleaning log.

    Args:
        inv: DataFrame de inventario limpio.
        orders: DataFrame de ordenes limpio.

    Returns:
        DataFrame con 3 columnas: Metric, Value, Detail.
    """
    inv_total = len(inv)
    inv_value = inv["Inventory Value"].sum()
    dead_n = int(inv["Dead Stock"].sum())
    dead_val = inv.loc[inv["Dead Stock"] == 1, "Inventory Value"].sum()
    over_n = int(inv["Overstock"].sum())
    over_val = inv.loc[inv["Overstock"] == 1, "Inventory Value"].sum()

    ord_total = len(orders)
    open_val = orders["Open Value"].sum()
    redun_n = int(orders["Redundant Order"].sum())
    redun_val = orders.loc[orders["Redundant Order"] == 1, "Open Value"].sum()
    old_n = int((orders["Age Days"] > 365).sum())

    pct = lambda n, total: f"{n / total * 100:.1f}%"
    usd = lambda v: f"${v:,.0f}"

    rows = [
        ("INVENTORY", "", ""),
        ("Total items", f"{inv_total:,}", ""),
        ("Inventory value", usd(inv_value), ""),
        ("Dead stock items", f"{dead_n:,}", pct(dead_n, inv_total)),
        ("Dead stock value", usd(dead_val), pct(dead_val, inv_value) + " of value"),
        ("Overstock items", f"{over_n:,}", pct(over_n, inv_total)),
        ("Overstock value", usd(over_val), pct(over_val, inv_value) + " of value"),
        ("", "", ""),
        ("OPEN ORDERS", "", ""),
        ("Total orders", f"{ord_total:,}", ""),
        ("Open value", usd(open_val), ""),
        ("Redundant orders", f"{redun_n:,}", pct(redun_n, ord_total)),
        ("Redundant value", usd(redun_val), "BOH >= Open Qty"),
        ("Orders older than 1 year", f"{old_n:,}", pct(old_n, ord_total)),
    ]

    return pd.DataFrame(rows, columns=["Metric", "Value", "Detail"])


def export_clean_excel(inv, orders, output_path=OUTPUT_PATH):
    """Exporta Excel limpio con 3 hojas formateadas.

    Hojas:
        - Summary: KPIs de inventario y ordenes.
        - Inventory: datos limpios + flags Dead Stock y Overstock.
        - Open Orders: datos limpios + Open Value, Redundant Order, Age Days.

    Args:
        inv: DataFrame de inventario limpio.
        orders: DataFrame de ordenes limpio.
        output_path: Ruta del archivo de salida.

    Returns:
        Ruta del archivo exportado.
    """
    summary = _build_summary(inv, orders)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        inv.to_excel(writer, sheet_name="Inventory", index=False)
        orders.to_excel(writer, sheet_name="Open Orders", index=False)

        wb = writer.book

        # Summary
        ws = wb["Summary"]
        _style_sheet(ws)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):
            if row[0].value in ("INVENTORY", "OPEN ORDERS"):
                row[0].font = SECTION_FONT

        # Inventory
        ws = wb["Inventory"]
        _style_sheet(ws)
        _apply_number_format(ws, inv, {
            "Current Price": CURRENCY_FMT,
            "Inventory Value": CURRENCY_FMT,
            "Balance On Hand": INTEGER_FMT,
            "Dead Stock": INTEGER_FMT,
            "Overstock": INTEGER_FMT,
        })

        # Open Orders
        ws = wb["Open Orders"]
        _style_sheet(ws)
        _apply_number_format(ws, orders, {
            "Price": CURRENCY_FMT,
            "Open Value": CURRENCY_FMT,
            "Order Qty": INTEGER_FMT,
            "Open Qty": INTEGER_FMT,
            "BOH": INTEGER_FMT,
            "Redundant Order": INTEGER_FMT,
            "Age Days": INTEGER_FMT,
        })

    return output_path
